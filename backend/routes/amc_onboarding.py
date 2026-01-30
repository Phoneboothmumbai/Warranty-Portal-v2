"""
AMC Onboarding Routes - Multi-step wizard API endpoints
"""
import io
import uuid
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends, Query, Request
from fastapi.responses import StreamingResponse
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from models.amc_onboarding import AMCOnboarding, AMCOnboardingUpdate
from utils.helpers import get_ist_isoformat
from services.auth import get_current_company_user, get_current_admin

router = APIRouter()

# Database will be injected
_db = None

def init_db(database):
    global _db
    _db = database


# ==================== COMPANY PORTAL ENDPOINTS ====================

@router.get("/portal/onboarding")
async def get_company_onboarding(user: dict = Depends(get_current_company_user)):
    """Get current onboarding for company (or create new draft)"""
    company_id = user.get("company_id")
    
    # Find existing onboarding (not converted)
    onboarding = await _db.amc_onboardings.find_one(
        {"company_id": company_id, "status": {"$ne": "converted"}},
        {"_id": 0}
    )
    
    if not onboarding:
        # Get company name
        company = await _db.companies.find_one({"id": company_id}, {"name": 1})
        
        # Create new draft
        new_onboarding = AMCOnboarding(
            company_id=company_id,
            company_name=company.get("name") if company else None,
            created_by=user.get("id")
        )
        await _db.amc_onboardings.insert_one(new_onboarding.model_dump())
        onboarding = new_onboarding.model_dump()
    
    return onboarding


@router.put("/portal/onboarding")
async def update_company_onboarding(data: AMCOnboardingUpdate, user: dict = Depends(get_current_company_user)):
    """Save draft or update onboarding (company can only edit draft or changes_requested)"""
    company_id = user.get("company_id")
    
    # Find existing onboarding
    onboarding = await _db.amc_onboardings.find_one(
        {"company_id": company_id, "status": {"$in": ["draft", "changes_requested"]}},
        {"_id": 0}
    )
    
    if not onboarding:
        raise HTTPException(status_code=404, detail="No editable onboarding found")
    
    # Build update
    update_data = {"updated_at": get_ist_isoformat()}
    
    for field, value in data.model_dump(exclude_none=True).items():
        update_data[field] = value
    
    await _db.amc_onboardings.update_one(
        {"id": onboarding["id"]},
        {"$set": update_data}
    )
    
    return await _db.amc_onboardings.find_one({"id": onboarding["id"]}, {"_id": 0})


@router.post("/portal/onboarding/submit")
async def submit_company_onboarding(user: dict = Depends(get_current_company_user)):
    """Submit onboarding for admin review"""
    company_id = user.get("company_id")
    
    # Find draft onboarding
    onboarding = await _db.amc_onboardings.find_one(
        {"company_id": company_id, "status": {"$in": ["draft", "changes_requested"]}},
        {"_id": 0}
    )
    
    if not onboarding:
        raise HTTPException(status_code=404, detail="No draft onboarding found")
    
    # Validate required fields (Step 8 confirmation)
    step8 = onboarding.get("step8_scope_confirmation", {})
    if not step8.get("information_accuracy_confirmed"):
        raise HTTPException(status_code=400, detail="Please confirm information accuracy before submitting")
    
    # Update status
    await _db.amc_onboardings.update_one(
        {"id": onboarding["id"]},
        {"$set": {
            "status": "submitted",
            "submitted_at": get_ist_isoformat(),
            "updated_at": get_ist_isoformat()
        }}
    )
    
    return {"message": "Onboarding submitted successfully", "status": "submitted"}


# ==================== ADMIN ENDPOINTS ====================

@router.get("/admin/onboardings")
async def list_onboardings(
    status: Optional[str] = None,
    admin: dict = Depends(get_current_admin)
):
    """List all onboarding submissions"""
    query = {}
    if status:
        query["status"] = status
    
    onboardings = await _db.amc_onboardings.find(query, {"_id": 0}).sort("updated_at", -1).to_list(500)
    
    # Enrich with company info
    for ob in onboardings:
        company = await _db.companies.find_one({"id": ob.get("company_id")}, {"name": 1, "contact_email": 1})
        if company:
            ob["company_name"] = company.get("name")
            ob["company_email"] = company.get("contact_email")
    
    return onboardings


@router.get("/admin/onboardings/{onboarding_id}")
async def get_onboarding_detail(onboarding_id: str, admin: dict = Depends(get_current_admin)):
    """Get detailed onboarding data"""
    onboarding = await _db.amc_onboardings.find_one({"id": onboarding_id}, {"_id": 0})
    if not onboarding:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    
    # Enrich with company info
    company = await _db.companies.find_one({"id": onboarding.get("company_id")}, {"_id": 0})
    onboarding["company"] = company
    
    return onboarding


@router.put("/admin/onboardings/{onboarding_id}")
async def update_onboarding_admin(
    onboarding_id: str,
    data: AMCOnboardingUpdate,
    admin: dict = Depends(get_current_admin)
):
    """Admin can edit any onboarding"""
    onboarding = await _db.amc_onboardings.find_one({"id": onboarding_id}, {"_id": 0})
    if not onboarding:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    
    update_data = {"updated_at": get_ist_isoformat()}
    for field, value in data.model_dump(exclude_none=True).items():
        update_data[field] = value
    
    await _db.amc_onboardings.update_one(
        {"id": onboarding_id},
        {"$set": update_data}
    )
    
    return await _db.amc_onboardings.find_one({"id": onboarding_id}, {"_id": 0})


@router.post("/admin/onboardings/{onboarding_id}/request-changes")
async def request_onboarding_changes(
    onboarding_id: str,
    data: dict,
    admin: dict = Depends(get_current_admin)
):
    """Request changes from company"""
    feedback = data.get("feedback", "")
    if not feedback:
        raise HTTPException(status_code=400, detail="Feedback is required")
    
    onboarding = await _db.amc_onboardings.find_one({"id": onboarding_id}, {"_id": 0})
    if not onboarding:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    
    await _db.amc_onboardings.update_one(
        {"id": onboarding_id},
        {"$set": {
            "status": "changes_requested",
            "admin_feedback": feedback,
            "feedback_given_at": get_ist_isoformat(),
            "feedback_given_by": admin.get("id"),
            "updated_at": get_ist_isoformat()
        }}
    )
    
    return {"message": "Changes requested", "status": "changes_requested"}


@router.post("/admin/onboardings/{onboarding_id}/approve")
async def approve_onboarding(
    onboarding_id: str,
    admin: dict = Depends(get_current_admin)
):
    """Approve onboarding"""
    onboarding = await _db.amc_onboardings.find_one({"id": onboarding_id}, {"_id": 0})
    if not onboarding:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    
    await _db.amc_onboardings.update_one(
        {"id": onboarding_id},
        {"$set": {
            "status": "approved",
            "approved_at": get_ist_isoformat(),
            "approved_by": admin.get("id"),
            "updated_at": get_ist_isoformat()
        }}
    )
    
    return {"message": "Onboarding approved", "status": "approved"}


@router.post("/admin/onboardings/{onboarding_id}/convert-to-amc")
async def convert_to_amc(
    onboarding_id: str,
    admin: dict = Depends(get_current_admin)
):
    """Convert approved onboarding to AMC contract and import devices"""
    onboarding = await _db.amc_onboardings.find_one({"id": onboarding_id}, {"_id": 0})
    if not onboarding:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    
    if onboarding.get("status") != "approved":
        raise HTTPException(status_code=400, detail="Only approved onboardings can be converted")
    
    # Extract data
    step1 = onboarding.get("step1_company_contract", {})
    company_id = onboarding.get("company_id")
    
    # Create AMC Contract
    from models.amc import AMCContract
    
    amc_contract = AMCContract(
        company_id=company_id,
        name=f"AMC - {onboarding.get('company_name', 'Unknown')}",
        amc_type="comprehensive",
        start_date=step1.get("amc_start_date", get_ist_isoformat()[:10]),
        end_date=step1.get("amc_end_date", get_ist_isoformat()[:10]),
        internal_notes=f"Created from onboarding {onboarding_id}"
    )
    
    await _db.amc_contracts.insert_one(amc_contract.model_dump())
    
    # Import devices from inventory
    step4 = onboarding.get("step4_device_inventory", {})
    devices = step4.get("devices", [])
    devices_created = 0
    
    for device in devices:
        device_doc = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "device_type": device.get("device_type", "Other"),
            "brand": device.get("brand"),
            "model": device.get("model"),
            "serial_number": device.get("serial_number"),
            "configuration": device.get("configuration"),
            "os_version": device.get("os_version"),
            "purchase_date": device.get("purchase_date"),
            "warranty_status": device.get("warranty_status"),
            "status": "active" if device.get("condition") == "working" else "maintenance",
            "location": device.get("physical_location"),
            "assigned_user": device.get("assigned_user"),
            "department": device.get("department"),
            "source": "onboarding",
            "onboarding_id": onboarding_id,
            "is_deleted": False,
            "created_at": get_ist_isoformat()
        }
        await _db.devices.insert_one(device_doc)
        devices_created += 1
    
    # Update onboarding status
    await _db.amc_onboardings.update_one(
        {"id": onboarding_id},
        {"$set": {
            "status": "converted",
            "converted_amc_id": amc_contract.id,
            "updated_at": get_ist_isoformat()
        }}
    )
    
    return {
        "message": "Onboarding converted to AMC",
        "amc_contract_id": amc_contract.id,
        "devices_created": devices_created
    }


# ==================== EXCEL TEMPLATE ====================

# Category to sheet mapping with sample data
CATEGORY_CONFIG = {
    "desktops": {
        "sheet_name": "Desktops",
        "device_types": ["Desktop"],
        "sample": ["Desktop", "Dell", "OptiPlex 7090", "ABC123XYZ", "i5-11500, 16GB RAM, 512GB SSD", 
                   "Windows 11 Pro", "2023-06-15", "Under OEM", "Working", "John Doe", "IT", "Floor 2, Desk 15"],
    },
    "laptops": {
        "sheet_name": "Laptops",
        "device_types": ["Laptop"],
        "sample": ["Laptop", "HP", "EliteBook 840 G8", "DEF456UVW", "i7-1165G7, 16GB RAM, 512GB SSD",
                   "Windows 11 Pro", "2022-03-20", "Extended", "Working", "Jane Smith", "Sales", "Mobile"],
    },
    "apple_devices": {
        "sheet_name": "Apple Devices",
        "device_types": ["Apple Mac", "Apple iPhone", "Apple iPad"],
        "sample": ["Apple Mac", "Apple", "MacBook Pro 14", "C02XL12345", "M3 Pro, 18GB RAM, 512GB SSD",
                   "macOS Sonoma 14.2", "2024-01-10", "Under OEM", "Working", "CEO", "Executive", "Corner Office"],
    },
    "servers": {
        "sheet_name": "Servers",
        "device_types": ["Server"],
        "sample": ["Server", "Dell", "PowerEdge R740", "JKL012MNO", "Xeon Gold 6248, 128GB RAM, 4TB RAID",
                   "Windows Server 2019", "2020-08-05", "Under OEM", "Working", "N/A", "IT", "Server Room"],
    },
    "network_devices": {
        "sheet_name": "Network Devices",
        "device_types": ["Router", "Switch", "Firewall"],
        "sample": ["Router", "Cisco", "ISR 4331", "FCZ2312A1BC", "3 GE ports, 2 NIM slots",
                   "IOS XE 17.3", "2021-06-01", "Under OEM", "Working", "N/A", "IT", "Server Room"],
    },
    "printers": {
        "sheet_name": "Printers & Scanners",
        "device_types": ["Printer", "Scanner"],
        "sample": ["Printer", "Canon", "imageRUNNER 2630i", "GHI789RST", "A3 MFP, 30ppm",
                   "N/A", "2021-01-10", "Expired", "Working", "Shared", "Admin", "Reception"],
    },
    "cctv": {
        "sheet_name": "CCTV & Access Control",
        "device_types": ["CCTV"],
        "sample": ["CCTV", "Hikvision", "DS-2CD2T86G2-4I", "HIK20241234", "8MP, 4mm lens, IR 80m",
                   "N/A", "2023-03-15", "Under OEM", "Working", "N/A", "Security", "Main Entrance"],
    },
    "wifi_aps": {
        "sheet_name": "Wi-Fi Access Points",
        "device_types": ["Access Point"],
        "sample": ["Access Point", "Ubiquiti", "U6-Pro", "24:5A:4C:AB:12:34", "Wi-Fi 6, 4x4 MIMO",
                   "N/A", "2022-11-20", "Under OEM", "Working", "N/A", "IT", "Conference Room"],
    },
    "ups": {
        "sheet_name": "UPS & Power Backup",
        "device_types": ["UPS"],
        "sample": ["UPS", "APC", "Smart-UPS 3000", "AS1234567890", "3000VA, 2700W, Line Interactive",
                   "N/A", "2021-09-01", "Extended", "Working", "N/A", "IT", "Server Room"],
    },
}


def _create_device_sheet(wb, sheet_name, device_types, sample_data):
    """Helper to create a formatted device inventory sheet"""
    ws = wb.create_sheet(sheet_name)
    
    # Headers
    headers = [
        "Device Type*", "Brand", "Model", "Serial Number*", "Configuration (RAM/Storage/CPU)",
        "OS Version", "Purchase Date (YYYY-MM-DD)", "Warranty Status", "Condition", 
        "Assigned User", "Department", "Physical Location"
    ]
    
    # Styles
    header_fill = PatternFill(start_color="0F62FE", end_color="0F62FE", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Sample data row
    for col_idx, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.border = thin_border
    
    # Add data validation for dropdowns
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Device Type dropdown (category-specific)
    device_types_str = f'"{",".join(device_types)}"'
    dv_device = DataValidation(type="list", formula1=device_types_str, showDropDown=False)
    dv_device.error = "Please select from the list"
    dv_device.errorTitle = "Invalid Device Type"
    ws.add_data_validation(dv_device)
    dv_device.add('A2:A1000')
    
    # Warranty Status dropdown
    warranty_statuses = '"Under OEM,Extended,Expired"'
    dv_warranty = DataValidation(type="list", formula1=warranty_statuses, showDropDown=False)
    ws.add_data_validation(dv_warranty)
    dv_warranty.add('H2:H1000')
    
    # Condition dropdown
    conditions = '"Working,Intermittent,Faulty"'
    dv_condition = DataValidation(type="list", formula1=conditions, showDropDown=False)
    ws.add_data_validation(dv_condition)
    dv_condition.add('I2:I1000')
    
    # Set column widths
    widths = [15, 12, 20, 18, 35, 18, 22, 15, 12, 18, 15, 20]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    return ws


@router.get("/portal/onboarding/device-template")
async def download_device_template(
    categories: Optional[str] = Query(None, description="Comma-separated category keys: desktops,laptops,apple_devices,servers,network_devices,printers,cctv,wifi_aps,ups")
):
    """Download Excel template for device inventory with category-specific tabs"""
    
    # Parse categories (if provided) or use all
    if categories:
        selected_categories = [c.strip() for c in categories.split(",") if c.strip() in CATEGORY_CONFIG]
    else:
        selected_categories = list(CATEGORY_CONFIG.keys())
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet (we'll create category-specific ones)
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create a sheet for each selected category
    for cat_key in selected_categories:
        config = CATEGORY_CONFIG[cat_key]
        _create_device_sheet(
            wb,
            config["sheet_name"],
            config["device_types"],
            config["sample"]
        )
    
    # If no categories selected, create a generic "All Devices" sheet
    if not selected_categories:
        ws = wb.create_sheet("All Devices")
        all_device_types = ["Desktop", "Laptop", "Apple Mac", "Apple iPhone", "Apple iPad", 
                           "Server", "Router", "Switch", "Firewall", "Printer", "Scanner", 
                           "CCTV", "Access Point", "UPS", "Other"]
        _create_device_sheet(
            wb,
            "All Devices",
            all_device_types,
            ["Desktop", "Dell", "OptiPlex 7090", "ABC123XYZ", "i5-11500, 16GB RAM, 512GB SSD", 
             "Windows 11 Pro", "2023-06-15", "Under OEM", "Working", "John Doe", "IT", "Floor 2, Desk 15"]
        )
    
    # Add instructions sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        "AMC DEVICE INVENTORY TEMPLATE",
        "",
        "Instructions:",
        "1. Fill in device details in the relevant category tabs",
        "2. Each tab is pre-configured for specific device types",
        "3. Fields marked with * are mandatory",
        "4. Use dropdowns where available for consistency",
        "5. Delete the sample rows before adding your data",
        "",
        "Tabs in this workbook:",
    ]
    
    # Add tab names to instructions
    for cat_key in selected_categories:
        config = CATEGORY_CONFIG[cat_key]
        instructions.append(f"- {config['sheet_name']}: {', '.join(config['device_types'])}")
    
    instructions.extend([
        "",
        "Warranty Status:",
        "- Under OEM: Currently under manufacturer warranty",
        "- Extended: Has extended/third-party warranty",
        "- Expired: No active warranty",
        "",
        "Condition:",
        "- Working: Fully functional",
        "- Intermittent: Has occasional issues",
        "- Faulty: Not working properly",
        "",
        "Configuration Format:",
        "For PCs/Laptops: 'CPU, RAM, Storage' (e.g., 'i5-11500, 16GB RAM, 512GB SSD')",
        "For Servers: Include RAID configuration",
        "For Network devices: Ports, modules",
        "For other devices: N/A or relevant specs",
    ])
    
    for row, text in enumerate(instructions, 1):
        cell = ws_info.cell(row=row, column=1, value=text)
        if row == 1:
            cell.font = Font(bold=True, size=14)
        elif text.endswith(":") and text != "":
            cell.font = Font(bold=True)
    
    ws_info.column_dimensions['A'].width = 70
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=AMC_Device_Inventory_Template.xlsx"}
    )


@router.post("/portal/onboarding/upload-devices")
async def upload_device_inventory(
    data: dict,
    user: dict = Depends(get_current_company_user)
):
    """Process uploaded device inventory from Excel (parsed on frontend)"""
    devices = data.get("devices", [])
    
    # Validate and normalize devices
    normalized_devices = []
    for device in devices:
        normalized = {
            "id": str(uuid.uuid4()),
            "device_type": device.get("Device Type*") or device.get("device_type", "Other"),
            "brand": device.get("Brand") or device.get("brand"),
            "model": device.get("Model") or device.get("model"),
            "serial_number": device.get("Serial Number*") or device.get("serial_number"),
            "configuration": device.get("Configuration (RAM/Storage/CPU)") or device.get("configuration"),
            "os_version": device.get("OS Version") or device.get("os_version"),
            "purchase_date": device.get("Purchase Date (YYYY-MM-DD)") or device.get("purchase_date"),
            "warranty_status": (device.get("Warranty Status") or device.get("warranty_status", "")).lower().replace(" ", "_"),
            "condition": (device.get("Condition") or device.get("condition", "")).lower(),
            "assigned_user": device.get("Assigned User") or device.get("assigned_user"),
            "department": device.get("Department") or device.get("department"),
            "physical_location": device.get("Physical Location") or device.get("physical_location"),
        }
        normalized_devices.append(normalized)
    
    return {"devices": normalized_devices, "count": len(normalized_devices)}
