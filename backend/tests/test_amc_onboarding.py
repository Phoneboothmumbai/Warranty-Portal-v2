"""
AMC Onboarding Wizard Backend Tests
Tests for multi-step wizard, device template generation, and Excel upload parsing
"""
import pytest
import requests
import os
import io
import openpyxl

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
COMPANY_USER_EMAIL = "jane@acme.com"
COMPANY_USER_PASSWORD = "company123"
ADMIN_EMAIL = "admin@demo.com"
ADMIN_PASSWORD = "admin123"


class TestDeviceTemplateAPI:
    """Tests for multi-sheet Excel template generation"""
    
    def test_device_template_no_categories(self):
        """Test template generation without categories - should return all device sheets"""
        response = requests.get(f"{BASE_URL}/api/portal/onboarding/device-template")
        assert response.status_code == 200
        assert response.headers.get('content-type') == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Parse Excel and verify sheets
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        # Should have all category sheets plus Instructions
        expected_sheets = ['Desktops', 'Laptops', 'Apple Devices', 'Servers', 'Network Devices', 
                          'Printers & Scanners', 'CCTV & Access Control', 'Wi-Fi Access Points', 
                          'UPS & Power Backup', 'Instructions']
        for sheet in expected_sheets:
            assert sheet in sheet_names, f"Missing sheet: {sheet}"
        print(f"✓ Template without categories has {len(sheet_names)} sheets: {sheet_names}")
    
    def test_device_template_selected_categories(self):
        """Test template generation with specific categories"""
        categories = "desktops,laptops,printers"
        response = requests.get(f"{BASE_URL}/api/portal/onboarding/device-template?categories={categories}")
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        # Should only have selected categories plus Instructions
        assert 'Desktops' in sheet_names
        assert 'Laptops' in sheet_names
        assert 'Printers & Scanners' in sheet_names
        assert 'Instructions' in sheet_names
        
        # Should NOT have unselected categories
        assert 'Servers' not in sheet_names
        assert 'CCTV & Access Control' not in sheet_names
        
        print(f"✓ Template with categories={categories} has sheets: {sheet_names}")
    
    def test_device_template_single_category(self):
        """Test template with single category"""
        response = requests.get(f"{BASE_URL}/api/portal/onboarding/device-template?categories=servers")
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        assert 'Servers' in sheet_names
        assert 'Instructions' in sheet_names
        assert len(sheet_names) == 2
        print(f"✓ Single category template has sheets: {sheet_names}")
    
    def test_device_template_sheet_headers(self):
        """Test that each sheet has correct headers"""
        response = requests.get(f"{BASE_URL}/api/portal/onboarding/device-template?categories=desktops")
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb['Desktops']
        
        # Check headers in row 1
        expected_headers = [
            "Device Type*", "Brand", "Model", "Serial Number*", "Configuration (RAM/Storage/CPU)",
            "OS Version", "Purchase Date (YYYY-MM-DD)", "Warranty Status", "Condition",
            "Assigned User", "Department", "Physical Location"
        ]
        
        for col, expected in enumerate(expected_headers, 1):
            actual = ws.cell(row=1, column=col).value
            assert actual == expected, f"Header mismatch at col {col}: expected '{expected}', got '{actual}'"
        
        print(f"✓ Sheet headers are correct")
    
    def test_device_template_sample_data(self):
        """Test that each sheet has sample data row"""
        response = requests.get(f"{BASE_URL}/api/portal/onboarding/device-template?categories=desktops")
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb['Desktops']
        
        # Check sample data in row 2
        device_type = ws.cell(row=2, column=1).value
        assert device_type == "Desktop"
        
        serial = ws.cell(row=2, column=4).value
        assert serial == "ABC123XYZ"  # Sample serial from CATEGORY_CONFIG
        
        print(f"✓ Sample data row present with device_type={device_type}, serial={serial}")


class TestCompanyOnboardingAPI:
    """Tests for company onboarding CRUD operations"""
    
    @pytest.fixture
    def company_token(self):
        """Get company user auth token"""
        response = requests.post(f"{BASE_URL}/api/company/auth/login", json={
            "email": COMPANY_USER_EMAIL,
            "password": COMPANY_USER_PASSWORD
        })
        if response.status_code != 200:
            pytest.skip(f"Company login failed: {response.status_code}")
        return response.json().get("token")
    
    def test_get_onboarding(self, company_token):
        """Test fetching onboarding data"""
        response = requests.get(
            f"{BASE_URL}/api/portal/onboarding",
            headers={"Authorization": f"Bearer {company_token}"}
        )
        assert response.status_code == 200
        
        data = response.json()
        assert "id" in data
        assert "status" in data
        assert "current_step" in data
        print(f"✓ GET onboarding: status={data.get('status')}, step={data.get('current_step')}")
    
    def test_update_onboarding_step3_categories(self, company_token):
        """Test updating device categories in Step 3"""
        # Update step 3 with device categories
        update_data = {
            "step3_device_categories": {
                "has_desktops": True,
                "has_laptops": True,
                "has_printers": True,
                "has_apple_devices": False,
                "has_servers": False,
                "has_network_devices": False,
                "has_cctv": False,
                "has_wifi_aps": False,
                "has_ups": False,
                "other_devices": ""
            }
        }
        
        response = requests.put(
            f"{BASE_URL}/api/portal/onboarding",
            headers={"Authorization": f"Bearer {company_token}"},
            json=update_data
        )
        assert response.status_code == 200
        
        data = response.json()
        step3 = data.get("step3_device_categories", {})
        assert step3.get("has_desktops") == True
        assert step3.get("has_laptops") == True
        assert step3.get("has_printers") == True
        print(f"✓ Step 3 categories updated: desktops={step3.get('has_desktops')}, laptops={step3.get('has_laptops')}, printers={step3.get('has_printers')}")
    
    def test_update_onboarding_step5_static_ip(self, company_token):
        """Test conditional field - has_static_ip and static_ip_addresses"""
        update_data = {
            "step5_network_infra": {
                "has_static_ip": True,
                "static_ip_addresses": "192.168.1.100, 192.168.1.101",
                "bandwidth": "100 Mbps",
                "router_firewall_brand": "Cisco",
                "router_firewall_model": "ISR 4331"
            }
        }
        
        response = requests.put(
            f"{BASE_URL}/api/portal/onboarding",
            headers={"Authorization": f"Bearer {company_token}"},
            json=update_data
        )
        assert response.status_code == 200
        
        data = response.json()
        step5 = data.get("step5_network_infra", {})
        assert step5.get("has_static_ip") == True
        assert step5.get("static_ip_addresses") == "192.168.1.100, 192.168.1.101"
        print(f"✓ Step 5 static IP: has_static_ip={step5.get('has_static_ip')}, addresses={step5.get('static_ip_addresses')}")
    
    def test_update_onboarding_step6_vpn_password_manager(self, company_token):
        """Test conditional fields - has_vpn/vpn_type and has_password_manager/password_manager_name"""
        update_data = {
            "step6_software_access": {
                "has_vpn": True,
                "vpn_type": "Cisco AnyConnect",
                "has_password_manager": True,
                "password_manager_name": "1Password",
                "email_platform": "google",
                "admin_access_available": True
            }
        }
        
        response = requests.put(
            f"{BASE_URL}/api/portal/onboarding",
            headers={"Authorization": f"Bearer {company_token}"},
            json=update_data
        )
        assert response.status_code == 200
        
        data = response.json()
        step6 = data.get("step6_software_access", {})
        assert step6.get("has_vpn") == True
        assert step6.get("vpn_type") == "Cisco AnyConnect"
        assert step6.get("has_password_manager") == True
        assert step6.get("password_manager_name") == "1Password"
        print(f"✓ Step 6 VPN: has_vpn={step6.get('has_vpn')}, vpn_type={step6.get('vpn_type')}")
        print(f"✓ Step 6 Password Manager: has_password_manager={step6.get('has_password_manager')}, name={step6.get('password_manager_name')}")
    
    def test_save_draft_persists_data(self, company_token):
        """Test that save draft persists data correctly"""
        # First update with test data
        test_company_name = "TEST_Draft_Company"
        update_data = {
            "current_step": 2,
            "step1_company_contract": {
                "company_name_legal": test_company_name,
                "brand_trade_name": "Test Brand",
                "gst_number": "GST123456789"
            }
        }
        
        response = requests.put(
            f"{BASE_URL}/api/portal/onboarding",
            headers={"Authorization": f"Bearer {company_token}"},
            json=update_data
        )
        assert response.status_code == 200
        
        # Fetch again to verify persistence
        get_response = requests.get(
            f"{BASE_URL}/api/portal/onboarding",
            headers={"Authorization": f"Bearer {company_token}"}
        )
        assert get_response.status_code == 200
        
        data = get_response.json()
        step1 = data.get("step1_company_contract", {})
        assert step1.get("company_name_legal") == test_company_name
        assert step1.get("gst_number") == "GST123456789"
        print(f"✓ Draft saved and persisted: company_name={step1.get('company_name_legal')}")


class TestDeviceUploadAPI:
    """Tests for device upload/import functionality"""
    
    @pytest.fixture
    def company_token(self):
        """Get company user auth token"""
        response = requests.post(f"{BASE_URL}/api/company/auth/login", json={
            "email": COMPANY_USER_EMAIL,
            "password": COMPANY_USER_PASSWORD
        })
        if response.status_code != 200:
            pytest.skip(f"Company login failed: {response.status_code}")
        return response.json().get("token")
    
    def test_upload_devices_normalizes_data(self, company_token):
        """Test that upload-devices endpoint normalizes Excel column names"""
        # Simulate parsed Excel data with Excel column names
        devices = [
            {
                "Device Type*": "Desktop",
                "Brand": "Dell",
                "Model": "OptiPlex 7090",
                "Serial Number*": "TEST_SN_001",
                "Configuration (RAM/Storage/CPU)": "i5, 16GB, 512GB",
                "OS Version": "Windows 11",
                "Purchase Date (YYYY-MM-DD)": "2024-01-15",
                "Warranty Status": "Under OEM",
                "Condition": "Working",
                "Assigned User": "Test User",
                "Department": "IT",
                "Physical Location": "Floor 1"
            },
            {
                "Device Type*": "Laptop",
                "Brand": "HP",
                "Model": "EliteBook 840",
                "Serial Number*": "TEST_SN_002",
                "Warranty Status": "Extended",
                "Condition": "Working"
            }
        ]
        
        response = requests.post(
            f"{BASE_URL}/api/portal/onboarding/upload-devices",
            headers={"Authorization": f"Bearer {company_token}"},
            json={"devices": devices}
        )
        assert response.status_code == 200
        
        data = response.json()
        assert data.get("count") == 2
        
        normalized = data.get("devices", [])
        assert len(normalized) == 2
        
        # Check first device normalization
        device1 = normalized[0]
        assert device1.get("device_type") == "Desktop"
        assert device1.get("serial_number") == "TEST_SN_001"
        assert device1.get("warranty_status") == "under_oem"  # Normalized to lowercase with underscore
        assert device1.get("condition") == "working"  # Normalized to lowercase
        assert "id" in device1  # UUID generated
        
        print(f"✓ Upload devices normalized {data.get('count')} devices")
        print(f"  Device 1: type={device1.get('device_type')}, serial={device1.get('serial_number')}")


class TestAdminOnboardingAPI:
    """Tests for admin onboarding management"""
    
    @pytest.fixture
    def admin_token(self):
        """Get admin auth token"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD
        })
        if response.status_code != 200:
            pytest.skip(f"Admin login failed: {response.status_code}")
        return response.json().get("token")
    
    def test_list_onboardings(self, admin_token):
        """Test admin can list all onboardings"""
        response = requests.get(
            f"{BASE_URL}/api/admin/onboardings",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        
        data = response.json()
        assert isinstance(data, list)
        print(f"✓ Admin list onboardings: {len(data)} onboardings found")
    
    def test_list_onboardings_filter_by_status(self, admin_token):
        """Test admin can filter onboardings by status"""
        response = requests.get(
            f"{BASE_URL}/api/admin/onboardings?status=draft",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        
        data = response.json()
        for ob in data:
            assert ob.get("status") == "draft"
        print(f"✓ Admin filter by status=draft: {len(data)} draft onboardings")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
